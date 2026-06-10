"""
slr/exporters/excel_exporter.py
=================================
Writes the article corpus to a styled Excel workbook (.xlsx).

Workbook structure
------------------
- **Summary** sheet (first): totals per source + grand total, run metadata
- One additional sheet per data source with its articles

Styling
-------
- Header rows: bold, white text on a dark teal background, auto-width columns
- Alternating row colours for readability
- Frozen top row in every sheet
- Hyperlinked DOI column (``https://doi.org/<doi>``)
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from slr.models import Article

logger = logging.getLogger(__name__)

_EXCEL_FILENAME = "articles_final.xlsx"

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")  # dark navy

# Alternating row fills
_ROW_FILL_EVEN = PatternFill(fill_type="solid", fgColor="DCE6F1")  # light blue
_ROW_FILL_ODD = PatternFill(fill_type="solid", fgColor="FFFFFF")   # white

_ARTICLE_COLUMNS = ["Title", "Authors", "Year", "DOI", "Source", "Abstract"]
_MIN_COL_WIDTH = 12
_MAX_COL_WIDTH = 80


def export_excel(articles: list[Article], output_dir: Path) -> Path:
    """Write *articles* to a styled Excel workbook.

    Returns the path to the written file.
    """
    output_path = output_dir / _EXCEL_FILENAME
    wb = Workbook()

    # Group articles by source
    by_source: dict[str, list[Article]] = defaultdict(list)
    for article in articles:
        by_source[article.source].append(article)

    # ── Summary sheet ───────────────────────────────────────────────────────
    _write_summary_sheet(wb, articles, by_source)

    # ── Per-source sheets ────────────────────────────────────────────────────
    for source_url, source_articles in sorted(by_source.items()):
        sheet_name = _source_to_sheet_name(source_url)
        ws = wb.create_sheet(title=sheet_name)
        _write_article_sheet(ws, source_articles)

    # Remove the default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    logger.info("Excel exported: %s (%d articles)", output_path.name, len(articles))
    return output_path


def _write_summary_sheet(
    wb: Workbook,
    all_articles: list[Article],
    by_source: dict[str, list[Article]],
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    # Title
    ws["A1"] = "SLR Pipeline — Run Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    ws["A2"] = f"Generated: {datetime.now(tz=timezone.utc).isoformat()} UTC"
    ws["A2"].font = Font(italic=True, size=10)

    ws["A3"] = f"Total articles (after dedup + DOI filter): {len(all_articles)}"
    ws["A3"].font = Font(bold=True, size=12)

    # Source breakdown table
    row = 5
    ws.cell(row=row, column=1, value="Source").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.cell(row=row, column=2, value="Article Count").font = _HEADER_FONT
    ws.cell(row=row, column=2).fill = _HEADER_FILL

    for i, (source, arts) in enumerate(sorted(by_source.items()), start=1):
        r = row + i
        fill = _ROW_FILL_EVEN if i % 2 == 0 else _ROW_FILL_ODD
        c1 = ws.cell(row=r, column=1, value=source)
        c1.fill = fill
        c2 = ws.cell(row=r, column=2, value=len(arts))
        c2.fill = fill

    # Auto-width
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18


def _write_article_sheet(ws, articles: list[Article]) -> None:
    """Populate a worksheet with articles and apply styling."""
    # Header row
    for col_idx, header in enumerate(_ARTICLE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, article in enumerate(articles, start=2):
        row_fill = _ROW_FILL_EVEN if row_idx % 2 == 0 else _ROW_FILL_ODD
        values = [
            article.title,
            article.authors_str,
            article.publication_year,
            article.doi,
            article.source,
            article.abstract,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=(col_idx in (1, 6)), vertical="top")

    # Add hyperlinks on the DOI column (col 4)
    for row_idx in range(2, len(articles) + 2):
        doi_cell = ws.cell(row=row_idx, column=4)
        doi_val = doi_cell.value or ""
        if doi_val and doi_val.startswith("10."):
            doi_cell.hyperlink = f"https://doi.org/{doi_val}"
            doi_cell.font = Font(color="0563C1", underline="single")

    # Auto-column widths
    _auto_fit_columns(ws)


def _auto_fit_columns(ws) -> None:  # type: ignore[no-untyped-def]
    """Set column widths based on the maximum cell content length."""
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_widths[cell.column] = min(
                    max(col_widths.get(cell.column, _MIN_COL_WIDTH), len(str(cell.value))),
                    _MAX_COL_WIDTH,
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _source_to_sheet_name(source_url: str) -> str:
    """Convert a source URL to a valid Excel sheet name (max 31 chars)."""
    name = (
        source_url.replace("https://", "")
        .replace("http://", "")
        .rstrip("/")
        .replace("www.", "")
    )
    # Excel sheet names: max 31 chars, no special chars
    for ch in r'\/[]*?:':
        name = name.replace(ch, "_")
    return name[:31]
