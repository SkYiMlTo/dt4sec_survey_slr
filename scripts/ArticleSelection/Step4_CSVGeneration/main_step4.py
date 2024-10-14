import os
import csv
import json
from openpyxl import Workbook


def excel_generation(input_path, output_path):
    fields = ['Title', 'Authors', 'Year of Publication', 'DOI', "Source"]

    # Get list of filenames from input directory
    input_filenames = next(os.walk(input_path), (None, None, []))[2]

    # Create a new Excel workbook
    workbook = Workbook()

    # Loop through each input file and add its content to a new sheet
    for input_filename in input_filenames:
        # Load JSON data from the input file
        file_input = json.load(open(input_path + input_filename, "r", encoding="utf-8"))

        # Create a new sheet in the Excel file with the name of the input file (without extension)
        sheet_name = os.path.splitext(input_filename)[0]
        sheet = workbook.create_sheet(title=sheet_name)

        # Write the header row
        sheet.append(fields)

        # Write each article's information into the sheet
        for item in file_input:
            title = item['title']
            authors = ', '.join([author['lastName'] for author in item['authors']]) if item['authors'] else 'No authors'
            publication_year = item['publicationYear']
            doi = item['doi']
            source = item['source']
            sheet.append([title, authors, publication_year, doi, source])

    # Remove the default sheet created by Workbook() if unused
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Save the Excel file to the output path
    workbook.save(output_path + "articles.xlsx")


def main_excel_generation(path):
    input_path = path + "step3/"
    output_path = path + "step4/"
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    excel_generation(input_path, output_path)
