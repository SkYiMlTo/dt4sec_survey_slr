import matplotlib.pyplot as plt
import os
import openpyxl

def years_repartition(input_path, output_path):
    publication_years = []

    # Load the Excel file
    workbook = openpyxl.load_workbook(input_path)

    # Loop through all sheets and collect publication years
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Skip the header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            publication_year = row[2]  # Assuming the publication year is in the 3rd column
            if publication_year:  # Check if the year exists
                publication_years.append(str(publication_year))  # Convert to string for consistency

    # Count the frequency of each year
    year_counts = {}
    for year in publication_years:
        if year in year_counts:
            year_counts[year] += 1
        else:
            year_counts[year] = 1

    # Sort years for plotting
    sorted_years = sorted(year_counts.keys())
    sorted_counts = [year_counts[year] for year in sorted_years]

    # Plotting
    plt.figure(figsize=(10, 5))
    plt.bar(sorted_years, sorted_counts, color='skyblue')
    plt.xlabel('Publication Year')
    plt.ylabel('Number of Publications')
    plt.title('Distribution of Publications by Year')
    plt.tight_layout()
    plt.savefig(output_path + "plt_years_repartition.png")
    plt.show()


def main_years_repartition(path):
    input_path = path + "articleSelection/step4/articles.xlsx"
    output_path = path + "dataVisualization/"
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    years_repartition(input_path, output_path)


# main_years_repartition("../../../output/2024-10-14_15-23-10.567515/")
