import openpyxl
import matplotlib.pyplot as plt
from collections import Counter
import os


def authors_contribution(input_path, output_path):
    # List to hold authors
    authors_list = []

    # Load the Excel file
    workbook = openpyxl.load_workbook(input_path)

    # Loop through all sheets to extract authors
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Skip the header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            authors = row[1]  # Assuming the authors are in the second column
            if authors:
                authors = authors.split(', ')  # Split the authors by comma
                if "No authors" in authors:
                    authors.remove("No authors")
                authors_list.extend(authors)

    # Count the frequency of each author
    author_counts = Counter(authors_list)

    # Get the top 10 authors with the most publications
    top_authors = author_counts.most_common(10)

    # Separate the authors and their counts for plotting
    authors, counts = zip(*top_authors)

    # Plotting
    plt.figure(figsize=(10, 5))
    plt.bar(authors, counts, color='skyblue')
    plt.xlabel('Authors')
    plt.ylabel('Number of Publications')
    plt.title('Top 10 Authors with Most Publications')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot as a file
    plot_filename = 'plt_top_10_authors.png'
    plt.savefig(output_path + plot_filename)


def main_authors_contribution(path):
    input_path = path + "articleSelection/step4/articles.xlsx"
    output_path = path + "dataVisualization/"
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    authors_contribution(input_path, output_path)


# main_authors_contribution("../../../output/2024-10-14_15-23-10.567515/")
